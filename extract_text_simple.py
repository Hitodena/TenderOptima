#!/usr/bin/env python3
"""
Simplified text extraction utility for SupplierFinder
Avoids numpy dependencies that cause import errors
"""

import sys
import os
from pathlib import Path

# Document processing imports - only essential ones
try:
    import PyPDF2
    from docx import Document
    import openpyxl
except ImportError as e:
    print(f"Missing required package: {e}", file=sys.stderr)
    sys.exit(1)

def extract_text_from_pdf(file_path):
    """Extract text from PDF using PyPDF2 only"""
    text = ""
    page_number = 1
    
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            for page in pdf_reader.pages:
                try:
                    page_text = page.extract_text()
                    if page_text.strip():
                        text += f"\n[PAGE:{page_number}]\n{page_text.strip()}\n"
                    page_number += 1
                except Exception as e:
                    print(f"Error extracting page {page_number}: {e}", file=sys.stderr)
                    page_number += 1
                    continue
                    
    except Exception as e:
        print(f"Error reading PDF: {e}", file=sys.stderr)
        return ""
    
    return text.strip()

def extract_text_from_docx(file_path):
    """Extract text from DOCX files with page markers"""
    text = ""
    page_number = 1
    
    try:
        doc = Document(file_path)
        current_page_text = ""
        
        for para in doc.paragraphs:
            para_text = para.text.strip()
            if para_text:
                current_page_text += para_text + "\n"
                
                # Add page break markers for long content
                if len(current_page_text) > 1500:
                    text += f"\n[PAGE:{page_number}]\n{current_page_text.strip()}\n"
                    current_page_text = ""
                    page_number += 1
        
        # Add remaining text
        if current_page_text.strip():
            text += f"\n[PAGE:{page_number}]\n{current_page_text.strip()}\n"
            
    except Exception as e:
        print(f"Error reading DOCX: {e}", file=sys.stderr)
        return ""
    
    return text.strip()

def extract_text_from_excel(file_path):
    """Extract text from Excel files"""
    text = ""
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n[SHEET:{sheet_name}]\n"
            
            for row in sheet.iter_rows(values_only=True):
                row_text = []
                for cell in row:
                    if cell is not None:
                        row_text.append(str(cell))
                if row_text:
                    text += " | ".join(row_text) + "\n"
                    
    except Exception as e:
        print(f"Error reading Excel: {e}", file=sys.stderr)
        return ""
    
    return text.strip()

def extract_text_from_txt(file_path):
    """Extract text from plain text files"""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            return file.read()
    except Exception as e:
        print(f"Error reading text file: {e}", file=sys.stderr)
        return ""

def main():
    if len(sys.argv) != 3:
        print("Usage: python3 extract_text_simple.py <file_path> <mime_type>", file=sys.stderr)
        sys.exit(1)
    
    file_path = sys.argv[1]
    mime_type = sys.argv[2]
    
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}", file=sys.stderr)
        sys.exit(1)
    
    # Determine file type and extract text
    file_extension = Path(file_path).suffix.lower()
    
    try:
        if file_extension == '.pdf' or 'pdf' in mime_type:
            text = extract_text_from_pdf(file_path)
        elif file_extension == '.docx' or 'wordprocessingml' in mime_type:
            text = extract_text_from_docx(file_path)
        elif file_extension in ['.xlsx', '.xls'] or 'spreadsheet' in mime_type:
            text = extract_text_from_excel(file_path)
        elif file_extension == '.txt' or 'text/plain' in mime_type:
            text = extract_text_from_txt(file_path)
        else:
            print(f"Unsupported file type: {file_extension}", file=sys.stderr)
            sys.exit(1)
        
        # Output the extracted text
        print(text)
        
    except Exception as e:
        print(f"Error extracting text: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()