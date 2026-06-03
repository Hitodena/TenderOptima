from pathlib import Path

import openpyxl

from backend.services.extraction.base import BaseExtractor, ExtractedDocument


class XlsxExtractor(BaseExtractor):
    @classmethod
    def extract(cls, file_path: Path) -> ExtractedDocument:
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        tables: list[str] = []

        for sheet in wb.worksheets:
            resolved = cls._resolve_merged_cells(sheet)
            rows = list(sheet.iter_rows(values_only=False))
            if not rows:
                continue
            md = cls._sheet_to_markdown(sheet.title, resolved, len(rows[0]))
            if md:
                tables.append(md)

        return ExtractedDocument(tables=tables)

    @classmethod
    def _resolve_merged_cells(cls, sheet) -> dict[tuple, str]:
        """Map every (row, col) → its effective value, filling merged cells."""
        resolved: dict[tuple, str] = {}

        for merge_range in sheet.merged_cells.ranges:
            top_left = sheet.cell(
                merge_range.min_row, merge_range.min_col
            ).value
            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    resolved[(row, col)] = str(top_left or "")

        for row in sheet.iter_rows():
            for cell in row:
                key = (cell.row, cell.column)
                if key not in resolved:
                    resolved[key] = str(cell.value or "").strip()

        return resolved

    @classmethod
    def _sheet_to_markdown(
        cls, title: str, resolved: dict[tuple, str], col_count: int
    ) -> str:
        if not resolved:
            return ""

        max_row = max(r for r, _ in resolved)
        lines = [f"**Лист: {title}**"]

        for row_idx in range(1, max_row + 1):
            cells = [
                resolved.get((row_idx, col), "")
                for col in range(1, col_count + 1)
            ]
            lines.append("| " + " | ".join(cells) + " |")
            if row_idx == 1:
                lines.append("| " + " | ".join(["---"] * col_count) + " |")

        return "\n".join(lines)
