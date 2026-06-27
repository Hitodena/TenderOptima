import io
import math
from typing import Any, Protocol

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from backend.enums import TZAnalysisStatus
from backend.schemas.analysis import TZAnalysisItem
from backend.services.analysis.docx_export import (
    _letter_tz_requirement_ref,
    _letter_tz_requirement_text,
)

_SUPPLIER_SUBCOLS = (
    "Первоначальное\nпредложение",
    "Последнее\nпредложение",
    "Статус",
    "Пояснение",
)
_HEADER_FILL = PatternFill("solid", fgColor="E8EEF4")
_CHANGED_FILL = PatternFill("solid", fgColor="FFF3CD")
_THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
_HEADER_FONT = Font(bold=True)
_STRIKE_FONT = Font(strike=True, color="666666")
_ALT_ROW_FILL = PatternFill("solid", fgColor="F8FAFC")
_TZ_STATUS_FILLS = {
    TZAnalysisStatus.MET: PatternFill("solid", fgColor="E8F5E9"),
    TZAnalysisStatus.PARTIAL: PatternFill("solid", fgColor="FFF8E1"),
    TZAnalysisStatus.MISSING: PatternFill("solid", fgColor="FFEBEE"),
    TZAnalysisStatus.NOT_FOUND: PatternFill("solid", fgColor="F1F3F5"),
}
_TZ_HEADER = (
    "№",
    "КП",
    "Требование ТЗ",
    "Ссылка ТЗ",
    "Предложение КП",
    "Ссылка КП",
    "Статус",
    "Объяснение",
)
_TZ_COL_WIDTHS: dict[int, tuple[float, float]] = {
    1: (6, 8),
    2: (14, 28),
    3: (24, 52),
    4: (18, 40),
    5: (24, 52),
    6: (18, 40),
    7: (14, 22),
    8: (20, 48),
}


class ComparisonSupplierLike(Protocol):
    company_name: str
    main_email: str
    values: dict[str, str | None]
    previous_values: dict[str, str | None]
    explanations: dict[str, str | None]
    corrected_from: dict[str, str | None]
    statuses: dict[str, str | None]


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Serialize an openpyxl workbook to XLSX bytes."""
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def write_sheet_rows(
    ws: Worksheet,
    header: list[str],
    rows: list[list[Any]],
) -> None:
    """Write header row and data rows to a worksheet."""
    ws.append(header)
    for row in rows:
        ws.append(row)


def _display_previous_value(
    supplier: ComparisonSupplierLike,
    req: str,
) -> str:
    corrected = supplier.corrected_from.get(req)
    current = supplier.values.get(req)
    if (
        corrected
        and str(corrected).strip()
        and current
        and str(current).strip()
        and corrected != current
    ):
        return str(corrected)
    prev = supplier.previous_values.get(req)
    return str(prev) if prev else ""


def _values_changed(previous: str, current: str) -> bool:
    prev = previous.strip()
    cur = current.strip()
    return bool(prev and cur and prev != cur)


def _style_header_cell(cell) -> None:
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    cell.border = _THIN_BORDER


def _style_data_cell(
    cell, *, changed: bool = False, strike: bool = False
) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = _THIN_BORDER
    if changed:
        cell.fill = _CHANGED_FILL
    if strike:
        cell.font = _STRIKE_FONT


def _cell_line_count(value: Any, col_width: float) -> int:
    if value is None:
        return 1
    text = str(value)
    explicit = text.count("\n") + 1
    chars_per_line = max(int(col_width * 1.05), 1)
    wrapped = max(
        (math.ceil(len(line) / chars_per_line) for line in text.split("\n")),
        default=1,
    )
    return max(explicit, wrapped)


def _auto_fit_columns(
    ws: Worksheet,
    *,
    col_specs: dict[int, tuple[float, float]],
    padding: float = 2.0,
) -> None:
    """Set column widths from content length with per-column min/max caps."""
    for col_idx in range(1, ws.max_column + 1):
        min_w, max_w = col_specs.get(col_idx, (10.0, 40.0))
        max_len = 0.0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            for line in str(value).split("\n"):
                max_len = max(max_len, float(len(line)))
        width = min(max_w, max(min_w, max_len * 1.05 + padding))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _auto_fit_row_heights(
    ws: Worksheet,
    *,
    start_row: int = 2,
    line_height: float = 15.0,
    min_height: float = 18.0,
    max_height: float = 120.0,
) -> None:
    """Adjust row heights for wrapped multi-line cell content."""
    for row_idx in range(start_row, ws.max_row + 1):
        max_lines = 1
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            letter = get_column_letter(col_idx)
            col_width = ws.column_dimensions[letter].width or 10.0
            max_lines = max(max_lines, _cell_line_count(cell.value, col_width))
        ws.row_dimensions[row_idx].height = min(
            max_height,
            max(min_height, max_lines * line_height),
        )


def build_tz_analysis_workbook(
    items: list[TZAnalysisItem],
    status_labels: dict[TZAnalysisStatus, str],
) -> Workbook:
    """Build a styled worksheet for TZ/KP analysis export."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return wb
    ws.title = "Анализ КП"

    for col_idx, title in enumerate(_TZ_HEADER, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        _style_header_cell(cell)

    for row_idx, item in enumerate(items, start=2):
        status_label = status_labels.get(item.status, item.status.value)
        values: list[Any] = [
            row_idx - 1,
            item.kp_name or "",
            _letter_tz_requirement_text(item),
            _letter_tz_requirement_ref(item) or (item.requirement_ref or ""),
            item.offer_value or "",
            item.offer_ref or "",
            status_label,
            item.explanation,
        ]
        alt_row = (row_idx - 2) % 2 == 1
        status_fill = _TZ_STATUS_FILLS.get(item.status)

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _style_data_cell(cell)
            if alt_row and col_idx != 7:
                cell.fill = _ALT_ROW_FILL
            if col_idx == 7 and status_fill is not None:
                cell.fill = status_fill
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
            if col_idx == 1:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="top",
                    wrap_text=True,
                )

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32
    _auto_fit_columns(ws, col_specs=_TZ_COL_WIDTHS)
    _auto_fit_row_heights(ws, start_row=2)
    return wb


def build_comparison_workbook(
    *,
    requirements: list[str],
    suppliers: list[ComparisonSupplierLike],
    status_labels: dict[str, str],
) -> Workbook:
    """Build a multi-row header comparison sheet with old/new value columns."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return wb
    ws.title = "Сравнение"

    cols_per_supplier = len(_SUPPLIER_SUBCOLS)
    ws.cell(row=1, column=1, value="Требование")
    _style_header_cell(ws.cell(row=1, column=1))
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    for idx, supplier in enumerate(suppliers):
        start_col = 2 + idx * cols_per_supplier
        end_col = start_col + cols_per_supplier - 1
        title = f"{supplier.company_name}\n{supplier.main_email}"
        ws.cell(row=1, column=start_col, value=title)
        _style_header_cell(ws.cell(row=1, column=start_col))
        if end_col > start_col:
            ws.merge_cells(
                start_row=1,
                start_column=start_col,
                end_row=1,
                end_column=end_col,
            )
            for col in range(start_col + 1, end_col + 1):
                _style_header_cell(ws.cell(row=1, column=col))
        for sub_idx, label in enumerate(_SUPPLIER_SUBCOLS):
            col = start_col + sub_idx
            ws.cell(row=2, column=col, value=label)
            _style_header_cell(ws.cell(row=2, column=col))

    for row_idx, req in enumerate(requirements, start=3):
        req_cell = ws.cell(row=row_idx, column=1, value=req)
        _style_data_cell(req_cell)
        req_cell.font = Font(bold=True)

        for sup_idx, supplier in enumerate(suppliers):
            start_col = 2 + sup_idx * cols_per_supplier
            previous = _display_previous_value(supplier, req)
            status_key = supplier.statuses.get(req)
            if status_key == "not_found":
                current = "—"
                status_label = ""
            else:
                current = supplier.values.get(req) or "—"
                status_label = (
                    status_labels.get(status_key, status_key)
                    if status_key
                    else ""
                )
            explanation = supplier.explanations.get(req) or ""
            changed = _values_changed(previous, current)

            prev_cell = ws.cell(row=row_idx, column=start_col, value=previous)
            _style_data_cell(prev_cell, changed=changed, strike=changed)

            cur_cell = ws.cell(
                row=row_idx,
                column=start_col + 1,
                value=current,
            )
            _style_data_cell(cur_cell, changed=changed)

            status_cell = ws.cell(
                row=row_idx,
                column=start_col + 2,
                value=status_label,
            )
            _style_data_cell(status_cell)

            expl_cell = ws.cell(
                row=row_idx,
                column=start_col + 3,
                value=explanation,
            )
            _style_data_cell(expl_cell)

    ws.freeze_panes = "B3"
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 42
    ws.column_dimensions["A"].width = 42
    for col_idx in range(
        2,
        2 + len(suppliers) * cols_per_supplier,
    ):
        letter = get_column_letter(col_idx)
        subcol_idx = (col_idx - 2) % cols_per_supplier
        if subcol_idx in (0, 1):
            ws.column_dimensions[letter].width = 24
        else:
            ws.column_dimensions[letter].width = 16

    return wb
